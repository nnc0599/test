from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from app.config import PRODUCT_CATEGORIES
from app.models.repositories import ProductRepository
from app.utils.product_prices import apply_normalized_product_prices
from app.utils.item_sort import sort_products_by_category_priority


PRODUCT_IMPORT_HEADERS = [
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Phân loại",
    "Đơn vị tính",
    "Số lượng",
    "Giá lẻ",
    "Giá thợ",
    "Giá đại lý",
    "Mô tả",
]

LEGACY_MULTI_PRICE_IMPORT_HEADERS = [
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Phân loại",
    "Đơn vị tính",
    "Số lượng",
    "Giá tùy chỉnh",
    "Giá lẻ",
    "Giá thợ",
    "Giá đại lý",
    "Mô tả",
]

LEGACY_PRODUCT_IMPORT_HEADERS = [
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Phân loại",
    "Đơn vị tính",
    "Số lượng",
    "Giá bán",
    "Mô tả",
]

LEGACY_PRODUCT_IMPORT_HEADERS_NO_CATEGORY = [
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Đơn vị tính",
    "Số lượng",
    "Giá bán",
    "Mô tả",
]

PRODUCT_EXPORT_HEADERS = [
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Phân loại",
    "Đơn vị tính",
    "Số lượng",
    "Giá lẻ",
    "Giá thợ",
    "Giá đại lý",
    "Mô tả",
    "Ghi chú",
    "Ngày cập nhật",
]

VALID_PRODUCT_CATEGORY_SET = set(PRODUCT_CATEGORIES)
VALID_PRODUCT_CATEGORIES_TEXT = ", ".join(PRODUCT_CATEGORIES)


class ProductController:
    def __init__(self, repo: ProductRepository):
        self.repo = repo

    def list_products(self) -> list[dict]:
        return sort_products_by_category_priority(self.repo.list_products())

    def search_products(self, keyword: str) -> list[dict]:
        if not keyword.strip():
            return sort_products_by_category_priority(self.repo.list_products())[:20]
        return sort_products_by_category_priority(self.repo.search_products(keyword))

    def create_product(self, payload: dict) -> None:
        if not payload.get("product_code", "").strip():
            raise ValueError("Mã sản phẩm không được để trống")
        if not payload.get("name", "").strip():
            raise ValueError("Tên sản phẩm không được để trống")
        payload = apply_normalized_product_prices(dict(payload))
        payload["category"] = self._normalize_category(payload.get("category"))
        if self.repo.get_product(payload["product_code"].strip()):
            raise ValueError("Mã sản phẩm đã tồn tại, vui lòng nhập mã khác")
        self.repo.add_product(payload)

    def parse_product_import_file(self, file_path: str) -> tuple[list[dict], int]:
        workbook = load_workbook(filename=Path(file_path), read_only=True, data_only=True)
        try:
            sheet = workbook.active

            header_row = [sheet[f"{column}1"].value for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]]
            normalized_headers = [str(value).strip() if value is not None else "" for value in header_row]
            import_mode = "new"
            if normalized_headers[: len(PRODUCT_IMPORT_HEADERS)] == PRODUCT_IMPORT_HEADERS:
                import_mode = "new"
            elif normalized_headers[: len(LEGACY_MULTI_PRICE_IMPORT_HEADERS)] == LEGACY_MULTI_PRICE_IMPORT_HEADERS:
                import_mode = "legacy_multi_price"
            elif normalized_headers[:7] == LEGACY_PRODUCT_IMPORT_HEADERS:
                import_mode = "legacy_with_category"
            elif normalized_headers[:6] == LEGACY_PRODUCT_IMPORT_HEADERS_NO_CATEGORY:
                import_mode = "legacy_no_category"
            else:
                raise ValueError("File không đúng biểu mẫu")

            products: list[dict] = []
            missing_code_count = 0
            if import_mode in {"new", "legacy_multi_price"}:
                max_col = 10
            elif import_mode == "legacy_with_category":
                max_col = 7
            else:
                max_col = 6
            for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
                if row is None:
                    continue
                values = list(row) + [None] * (max_col - len(row))
                if all(value is None or str(value).strip() == "" for value in values):
                    continue

                product_code = self._as_text(values[0])
                if not product_code:
                    missing_code_count += 1
                    continue

                row_number = len(products) + missing_code_count + 2
                if import_mode == "new":
                    payload = {
                        "product_code": product_code,
                        "name": self._as_text(values[1]),
                        "category": self._normalize_category(values[2], required=True, row_number=row_number),
                        "unit": self._as_text(values[3]),
                        "quantity": self._as_int(values[4]),
                        "retail_price": self._as_int(values[5]),
                        "worker_price": self._as_int(values[6]),
                        "dealer_price": self._as_int(values[7]),
                        "description": self._as_text(values[8]),
                        "note": "",
                    }
                elif import_mode == "legacy_multi_price":
                    retail_price = self._as_int(values[6]) or self._as_int(values[5])
                    payload = {
                        "product_code": product_code,
                        "name": self._as_text(values[1]),
                        "category": self._normalize_category(values[2], required=True, row_number=row_number),
                        "unit": self._as_text(values[3]),
                        "quantity": self._as_int(values[4]),
                        "retail_price": retail_price,
                        "worker_price": self._as_int(values[7]),
                        "dealer_price": self._as_int(values[8]),
                        "description": self._as_text(values[9]),
                        "note": "",
                    }
                elif import_mode == "legacy_with_category":
                    payload = {
                        "product_code": product_code,
                        "name": self._as_text(values[1]),
                        "category": self._normalize_category(values[2], required=True, row_number=row_number),
                        "unit": self._as_text(values[3]),
                        "quantity": self._as_int(values[4]),
                        "retail_price": self._as_int(values[5]),
                        "description": self._as_text(values[6]),
                        "note": "",
                    }
                else:
                    payload = {
                        "product_code": product_code,
                        "name": self._as_text(values[1]),
                        "category": "",
                        "unit": self._as_text(values[2]),
                        "quantity": self._as_int(values[3]),
                        "retail_price": self._as_int(values[4]),
                        "description": self._as_text(values[5]),
                        "note": "",
                    }

                products.append(apply_normalized_product_prices(payload))

            return products, missing_code_count
        finally:
            workbook.close()

    def create_products(self, payloads: list[dict]) -> None:
        if not payloads:
            return
        normalized_payloads = []
        for payload in payloads:
            item = apply_normalized_product_prices(dict(payload))
            item["category"] = self._normalize_category(item.get("category"), required=False)
            normalized_payloads.append(item)
        self.repo.add_products(normalized_payloads)

    def export_product_import_template(self, file_path: str) -> None:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "SanPham"
            sheet.append(PRODUCT_IMPORT_HEADERS)
            sheet.append([
                "SP001",
                "Tên sản phẩm mẫu",
                "Pin NLMT",
                "cái",
                0,
                1250000,
                1180000,
                1120000,
                "Mô tả sản phẩm mẫu",
            ])
            sheet.append([
                "SP002",
                "Sản phẩm chỉ có 1 giá",
                "Phụ kiện",
                "cái",
                0,
                45000,
                0,
                0,
                "Để trống giá thợ/đại lý nếu không áp dụng",
            ])
            workbook.save(file_path)
        finally:
            workbook.close()

    def export_products_to_excel(self, file_path: str) -> None:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "SanPham"
            sheet.append(PRODUCT_EXPORT_HEADERS)

            for product in self.list_products():
                sheet.append(
                    [
                        self._as_text(product.get("product_code")),
                        self._as_text(product.get("name")),
                        self._as_text(product.get("category")),
                        self._as_text(product.get("unit")),
                        self._as_int(product.get("quantity")),
                        self._as_int(product.get("retail_price")),
                        self._as_int(product.get("worker_price")),
                        self._as_int(product.get("dealer_price")),
                        self._as_text(product.get("description")),
                        self._as_text(product.get("note")),
                        self._as_text(product.get("updated_at")),
                    ]
                )

            workbook.save(file_path)
        finally:
            workbook.close()

    @staticmethod
    def _as_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @staticmethod
    def _as_int(value: object) -> int:
        if value is None:
            return 0
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)

        text = str(value).strip()
        if not text:
            return 0

        try:
            return int(text)
        except ValueError:
            return int(float(text))

    def update_product(self, product_code: str, payload: dict) -> None:
        if not payload.get("note", "").strip():
            raise ValueError("Ghi chú là bắt buộc khi sửa thông tin sản phẩm")
        payload = apply_normalized_product_prices(dict(payload))
        payload["category"] = self._normalize_category(payload.get("category"))
        self.repo.update_product(product_code, payload)

    @staticmethod
    def _normalize_category(
        value: object,
        *,
        required: bool = True,
        row_number: int | None = None,
    ) -> str:
        text = ProductController._as_text(value)
        if not text:
            if required:
                if row_number is not None:
                    raise ValueError(f"Thiếu phân loại tại dòng {row_number}")
                raise ValueError("Phân loại sản phẩm không được để trống")
            return ""
        if text not in VALID_PRODUCT_CATEGORY_SET:
            if row_number is not None:
                raise ValueError(
                    f"Phân loại không hợp lệ tại dòng {row_number}. Chỉ chấp nhận: {VALID_PRODUCT_CATEGORIES_TEXT}"
                )
            raise ValueError(
                f"Phân loại không hợp lệ. Chỉ chấp nhận: {VALID_PRODUCT_CATEGORIES_TEXT}"
            )
        return text

    def list_product_update_history(self, product_code: str) -> list[dict]:
        return self.repo.list_product_update_history(product_code)

    def delete_product(self, product_code: str) -> None:
        self.repo.delete_product(product_code)
